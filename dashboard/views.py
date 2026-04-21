import json
import statistics
import openpyxl
from io import BytesIO
from datetime import datetime
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt


# ── helpers ───────────────────────────────────────────────────────────

def to_str(val):
    return str(val).strip() if val is not None else ''

def parse_time(val):
    if val is None:
        return ''
    if hasattr(val, 'strftime'):
        return val.strftime('%H:%M:%S')
    return str(val).strip()

def parse_date(val, fallback=''):
    if val is None:
        return fallback
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d')
    return str(val).strip()

def get_dow(date_str):
    try:
        return datetime.strptime(date_str[:10], '%Y-%m-%d').weekday()
    except Exception:
        return -1

def get_hour(time_str):
    try:
        return int(str(time_str)[:2])
    except Exception:
        return -1

def is_refill(status):
    return 'refill' in status.lower()

def is_battery(status):
    return 'batter' in status.lower()


# ── amp limits ────────────────────────────────────────────────────────

def compute_amp_limits(rows, expected_current=None, tolerance_pct=10):
    if expected_current and expected_current > 0:
        t  = tolerance_pct / 100.0
        lo = round(expected_current * (1 - t), 2)
        hi = round(expected_current * (1 + t), 2)
        return lo, hi
    amps = [r['amp'] for r in rows if r['amp'] > 0]
    if len(amps) < 2:
        return 0, 0
    mean = sum(amps) / len(amps)
    std  = statistics.stdev(amps)
    return round(mean - 2 * std, 2), round(mean + 2 * std, 2)


def compute_accuracy(rows, lo, hi):
    """Percentage of amp readings within [lo, hi]."""
    if not rows:
        return 0
    within = sum(1 for r in rows if lo <= r['amp'] <= hi)
    return round(within / len(rows) * 100, 1)


def compute_within_range(rows, lo, hi):
    """
    COUNT of readings where amp >= lo AND amp <= hi
    i.e. reading is within BOTH lower and upper limit
    """
    return sum(1 for r in rows if lo <= r['amp'] <= hi)


def compute_out_of_range(rows, lo, hi):
    """
    Count readings OUTSIDE [lo, hi]
    i.e. amp < lo OR amp > hi
    """
    return sum(1 for r in rows if r['amp'] < lo or r['amp'] > hi)


# ── parser ────────────────────────────────────────────────────────────

def parse_workbook(file):
    wb       = openpyxl.load_workbook(file, data_only=True)
    all_rows = []
    sheets   = []
    summary  = {}

    for sheet_name in wb.sheetnames:
        ws        = wb[sheet_name]
        rows_iter = list(ws.iter_rows(values_only=True))
        if not rows_iter:
            continue
        headers = [to_str(h) for h in rows_iter[0]]

        if 'Glucose' not in headers:
            for row in rows_iter[1:]:
                if not row or row[0] is None:
                    continue
                k = to_str(row[0])
                v = to_str(row[1]) if len(row) > 1 and row[1] is not None else ''
                if k:
                    summary[k] = v
            continue

        sheets.append(sheet_name)
        status_col_idx = len(headers) - 1

        for row in rows_iter[1:]:
            if not row or all(v is None for v in row):
                continue
            record   = dict(zip(headers, row))
            glucose  = record.get('Glucose')
            amp      = record.get('Amp')
            status   = to_str(row[status_col_idx]) if len(row) > status_col_idx else ''
            date_str = parse_date(record.get('Date'), sheet_name)
            time_str = parse_time(record.get('Time'))

            all_rows.append({
                'id':       to_str(record.get('ID')),
                'userId':   to_str(record.get('User ID')),
                'glucose':  float(glucose) if glucose is not None else 0,
                'amp':      float(amp)     if amp     is not None else 0,
                'date':     date_str,
                'time':     time_str,
                'note':     to_str(record.get('Note')),
                'deviceId': to_str(record.get('Device ID')),
                'status':   status,
                'sheet':    sheet_name,
                'dow':      get_dow(date_str),
                'hour':     get_hour(time_str),
            })

    return all_rows, sheets, summary


# ── cleaning ──────────────────────────────────────────────────────────

def clean_rows(rows, apply_time_filter, amp_lo=0, amp_hi=9999):
    """
    Rules:
      1. Remove Sat (dow=5) and Sun (dow=6)
      2. If apply_time_filter: keep ONLY hour 10 to 18 inclusive
         FIX: 10am to 7pm means last allowed hour = 18 (up to 18:59)
         hour 19 = 7:00pm onwards → EXCLUDED
      3. For every Refilled/battery row:
           - Check if refill row itself is stable (amp within [amp_lo, amp_hi])
               * Stable  → KEEP refill row
               * Unstable → REMOVE refill row too
           - For each of the 3 rows BEFORE and 3 rows AFTER:
               * If amp is OUTSIDE [amp_lo, amp_hi] → DROP (unstable)
               * If amp is WITHIN  [amp_lo, amp_hi] → KEEP (already stable)
    """
    # Step 1 — remove weekends
    rows = [r for r in rows if r['dow'] not in (5, 6)]

    # Step 2 — time filter
    # FIX: 10am to 7pm = hour 10,11,12,13,14,15,16,17,18 only
    # 7pm = 19:00 → excluded (client does NOT want 7pm entries)
    if apply_time_filter:
        rows = [r for r in rows if 10 <= r['hour'] <= 18]

    # Step 3 — smart refill removal
    refill_markers = []
    drop_positions = set()

    for i, r in enumerate(rows):
        st = r['status']
        if is_refill(st) or is_battery(st):

            # FIX 3: Check if refill row itself is stable
            refill_amp = r['amp']
            refill_stable = (amp_lo <= refill_amp <= amp_hi)

            if refill_stable:
                # Refill row is within limits → KEEP it as marker
                refill_markers.append({
                    'date':      r['date'],
                    'time':      r['time'],
                    'timestamp': r['date'] + ' ' + r['time'],
                    'amp':       refill_amp,
                    'glucose':   r['glucose'],
                    'label':     st,
                    'sheet':     r['sheet'],
                    'stable':    True,
                })
                # Refill row itself is NOT dropped (stable)
            else:
                # Refill row is outside limits → REMOVE it too
                refill_markers.append({
                    'date':      r['date'],
                    'time':      r['time'],
                    'timestamp': r['date'] + ' ' + r['time'],
                    'amp':       refill_amp,
                    'glucose':   r['glucose'],
                    'label':     st,
                    'sheet':     r['sheet'],
                    'stable':    False,
                })
                drop_positions.add(i)  # drop the unstable refill row itself

            # Check 3 rows BEFORE the refill
            for j in range(max(0, i - 3), i):
                if not (amp_lo <= rows[j]['amp'] <= amp_hi):
                    drop_positions.add(j)   # unstable → DROP
                # else stable → KEEP

            # Check 3 rows AFTER the refill
            for j in range(i + 1, min(len(rows), i + 4)):
                if not (amp_lo <= rows[j]['amp'] <= amp_hi):
                    drop_positions.add(j)   # unstable → DROP
                # else stable → KEEP

    cleaned = [r for i, r in enumerate(rows) if i not in drop_positions]
    return cleaned, refill_markers


# ── views ─────────────────────────────────────────────────────────────

def index(request):
    return render(request, 'dashboard/index.html')


@csrf_exempt
def upload_file(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    f = request.FILES.get('file')
    if not f:
        return JsonResponse({'error': 'No file'}, status=400)

    try:
        raw_rows, sheets, summary = parse_workbook(f)

        # Fallback limits from raw data for initial cleaning
        lo_raw, hi_raw = compute_amp_limits(raw_rows)

        c1, m1 = clean_rows(raw_rows, apply_time_filter=True,
                             amp_lo=lo_raw, amp_hi=hi_raw)
        lo1, hi1 = compute_amp_limits(c1)

        c2, m2 = clean_rows(raw_rows, apply_time_filter=False,
                             amp_lo=lo_raw, amp_hi=hi_raw)
        lo2, hi2 = compute_amp_limits(c2)

        return JsonResponse({
            'success': True,
            'rawRows': raw_rows,
            'sheets':  sheets,
            'summary': summary,
            'v1':  {'rows': c1,       'refillMarkers': m1, 'ampLower': lo1,    'ampUpper': hi1},
            'v2':  {'rows': c2,       'refillMarkers': m2, 'ampLower': lo2,    'ampUpper': hi2},
            'raw': {'rows': raw_rows, 'refillMarkers': [], 'ampLower': lo_raw, 'ampUpper': hi_raw},
        })
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'trace': traceback.format_exc()}, status=500)


@csrf_exempt
def filter_data(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    try:
        body       = json.loads(request.body)
        rows       = body.get('rows', [])
        filters    = body.get('filters', {})
        markers_in = body.get('refillMarkers', [])

        expected_current = body.get('expectedCurrent', None)
        tolerance_pct    = body.get('tolerancePct', 10)

        sheet       = filters.get('sheet', 'all')
        user_id     = filters.get('userId', 'all')
        statuses    = filters.get('statuses', [])
        glucose_min = float(filters.get('glucoseMin', 0))
        glucose_max = float(filters.get('glucoseMax', 9999))
        amp_max     = float(filters.get('ampMax', 9999))
        time_from   = filters.get('timeFrom', '00:00')
        time_to     = filters.get('timeTo', '23:59')

        result = []
        for r in rows:
            if sheet   != 'all' and r.get('sheet')  != sheet:   continue
            if user_id != 'all' and r.get('userId') != user_id: continue
            if statuses and r.get('status') not in statuses:     continue
            g = r.get('glucose', 0)
            if g < glucose_min or g > glucose_max:               continue
            if r.get('amp', 0) > amp_max:                        continue
            t = (r.get('time') or '')[:5]
            if t and (t < time_from or t > time_to):             continue
            result.append(r)

        markers_out = markers_in
        if sheet != 'all':
            markers_out = [m for m in markers_in if m.get('sheet', '') == sheet]

        # Compute limits first — needed for smart refill drop
        ec = float(expected_current) if expected_current else None
        lo, hi = compute_amp_limits(result, expected_current=ec,
                                    tolerance_pct=float(tolerance_pct))

        # Re-apply smart refill drop with current limits
        # Unstable refill rows (amp outside [lo,hi]) are removed from result
        # but their markers are already in markers_out for chart display
        drop_pos = set()
        for i, r in enumerate(result):
            st = r.get('status', '')
            if is_refill(st) or is_battery(st):
                # Remove unstable refill row itself
                if not (lo <= r['amp'] <= hi):
                    drop_pos.add(i)
                # Remove unstable rows in ±3 window
                for j in range(max(0, i - 3), i):
                    if not (lo <= result[j]['amp'] <= hi):
                        drop_pos.add(j)
                for j in range(i + 1, min(len(result), i + 4)):
                    if not (lo <= result[j]['amp'] <= hi):
                        drop_pos.add(j)

        result = [r for i, r in enumerate(result) if i not in drop_pos]

        accuracy     = compute_accuracy(result, lo, hi)
        within_range = compute_within_range(result, lo, hi)
        out_of_range = compute_out_of_range(result, lo, hi)

        return JsonResponse({
            'rows':          result,
            'total':         len(result),
            'refillMarkers': markers_out,
            'ampLower':      lo,
            'ampUpper':      hi,
            'accuracy':      accuracy,
            'withinRange':   within_range,
            'outOfRange':    out_of_range,
        })
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'trace': traceback.format_exc()}, status=500)


@csrf_exempt
def export_excel(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    try:
        body             = json.loads(request.body)
        rows             = body.get('rows', [])
        amp_lo           = body.get('ampLower', 0)
        amp_hi           = body.get('ampUpper', 0)
        ds_label         = body.get('label', 'Clean data')
        solution         = body.get('solution', '—')
        expected_current = body.get('expectedCurrent', '—')
        tolerance_pct    = body.get('tolerancePct', 10)
        accuracy         = body.get('accuracy', '—')
        out_of_range     = body.get('outOfRange', '—')

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        from collections import defaultdict
        from openpyxl.styles import Font, PatternFill, Alignment

        groups = defaultdict(list)
        for r in rows:
            groups[r.get('sheet', r.get('date', 'Unknown'))].append(r)

        HEADERS = ['ID', 'User ID', 'Glucose', 'Amp',
                   'Lower Limit', 'Upper Limit',
                   'Within Range', 'Date', 'Time', 'Note', 'Device ID', 'Status']

        hfill    = PatternFill('solid', start_color='1D9E75', end_color='1D9E75')
        ok_fill  = PatternFill('solid', start_color='C6EFCE', end_color='C6EFCE')
        bad_fill = PatternFill('solid', start_color='FFC7CE', end_color='FFC7CE')

        for sheet_name, sheet_rows in groups.items():
            ws = wb.create_sheet(title=sheet_name[:31])
            ws.append(HEADERS)
            for cell in ws[1]:
                cell.font      = Font(bold=True, color='FFFFFF', size=11)
                cell.fill      = hfill
                cell.alignment = Alignment(horizontal='center')

            for r in sheet_rows:
                amp    = r.get('amp', 0)
                within = 'YES' if amp_lo <= amp <= amp_hi else 'NO'
                ws.append([
                    r.get('id', ''),      r.get('userId', ''),
                    r.get('glucose', ''), amp,
                    amp_lo,               amp_hi,
                    within,
                    r.get('date', ''),    r.get('time', ''),
                    r.get('note', ''),    r.get('deviceId', ''),
                    r.get('status', ''),
                ])
                cell      = ws.cell(row=ws.max_row, column=7)
                cell.fill = ok_fill if within == 'YES' else bad_fill
                cell.font = Font(bold=True,
                                 color='006100' if within == 'YES' else '9C0006')

            for col in ws.columns:
                w = max((len(str(c.value or '')) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(w + 4, 40)

        ws_sum = wb.create_sheet(title='Summary', index=0)
        sf = PatternFill('solid', start_color='1D9E75', end_color='1D9E75')
        summary_rows = [
            ['Key', 'Value'],
            ['Dataset',                          ds_label],
            ['Solution',                         solution],
            ['Expected Current (nA)',             expected_current],
            ['Tolerance (%)',                     f'{tolerance_pct}%'],
            ['Min Range (Lower Limit)',           amp_lo],
            ['Max Range (Upper Limit)',           amp_hi],
            ['Total Readings',                   len(rows)],
            ['Out of Range Readings',             out_of_range],
            ['Accuracy (within tolerance)',       f'{accuracy}%'],
            ['Sheets Included',                  ', '.join(groups.keys())],
        ]
        for i, row in enumerate(summary_rows, start=1):
            ws_sum.append(row)
            if i == 1:
                for cell in ws_sum[1]:
                    cell.font = Font(bold=True, color='FFFFFF', size=11)
                    cell.fill = sf
        ws_sum.column_dimensions['A'].width = 40
        ws_sum.column_dimensions['B'].width = 28

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = 'attachment; filename="glucose_clean_export.xlsx"'
        return resp

    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'trace': traceback.format_exc()}, status=500)